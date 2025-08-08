import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select

# Excel Utility:
def get_row_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    return sheet.max_row


def get_col_count(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    return sheet.max_column


def get_cell_data_by_row_col(file_path, sheet_name, row_num, col_num):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    return sheet.cell(row=row_num, column=col_num).value


def get_cell_data_by_column_name(file_path, sheet_name, row_num, column_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    for cols in range(1, sheet.max_column):
        if sheet.cell(row=1, column=cols).value == column_name:
            return sheet.cell(row=row_num, column=cols).value


def write_cell_data(file_path, sheet_name, row_num, col_num, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    sheet.cell(row=row_num, column=col_num).value = data
    workbook.save(file_path)


file = "..\\excel_data\\testdata.xlsx"
sheet = "Login_Test"

print(get_row_count(file, sheet))
print(get_col_count(file, sheet))

print(get_cell_data_by_row_col(file, sheet, 2, 2))
print(get_cell_data_by_column_name(file, sheet, 3, "password"))
write_cell_data(file, sheet, 1, 4, "DOB")
