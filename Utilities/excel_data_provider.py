import openpyxl


def get_data(sheet_name):
    workbook = openpyxl.load_workbook("..\\excel\\testdata.xlsx")
    sheet = workbook[sheet_name]
    main_list = []

    for i in range(2, sheet.max_row + 1):
        data_list = []
        for j in range(1, sheet.max_column + 1):
            data = sheet.cell(row=i, column=j).value
            data_list.insert(j, data)

        main_list.insert(i, data_list)

    return main_list  # This will return data directly from excel sheet not hardcoded data.